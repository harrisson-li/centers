# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import csv
import time
from openpyxl import Workbook, load_workbook


class CentersPipeline(object):
    def process_item(self, item, spider):
        file_name = 'school_number.xlsx'
        sheet_name = time.strftime("%Y-%m-%d")

        data = [item['brand'], item['name'], item['number']] + item['center']

        try:
            wb = load_workbook(file_name)
        except FileNotFoundError:
            wb = Workbook()
            ws = wb['Sheet']
            ws['A1'] = 'Check school number by date and brand'
            ws['A2'] = 'You need to sort by brand manually'
            wb.save(file_name)
            wb = load_workbook(file_name)

        if sheet_name not in wb.get_sheet_names():
            sheet = wb.create_sheet(sheet_name)
        else:
            sheet = wb[sheet_name]



        # if sheet_name not in wb.get_sheet_names():
        #     sheet = wb.create_sheet(sheet_name)
        # else:
        #     i = 1
        #     while (sheet_name + '-{}'.format(i)) in wb.get_sheet_names():
        #         i = i + 1
        #     sheet = wb.create_sheet(sheet_name + '-{}'.format(i))

        sheet.append(data)

        wb.save(file_name)


