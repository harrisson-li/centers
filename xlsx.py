from openpyxl import Workbook, load_workbook
import time


if __name__ == '__main__':
    file_name = 'school_number_1.xlsx'
    sheet_name = time.strftime("%Y-%m-%d")

    try:
        wb = load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(file_name)
        wb = load_workbook(file_name)

    if sheet_name not in wb.get_sheet_names():
        sheet = wb.create_sheet(sheet_name)
    else:
        sheet = wb[sheet_name]

    sheet.append([1, 2])
    print(sheet.max_row)
    print(sheet.max_column)
    sheet.

    wb.save(file_name)
